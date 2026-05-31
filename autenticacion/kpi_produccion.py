from __future__ import annotations

from datetime import date, datetime, timedelta
from decimal import Decimal, ROUND_HALF_UP
from io import BytesIO

from django.db.models import Prefetch, Q
from django.utils import timezone

from .models import (
    BOMOperacion,
    CostoHoraMaquina,
    CostoHoraOperador,
    InformeValidacionDefectoQA,
    LoteProduccion,
    OrdenFabricacion,
    OrdenFabricacionDetalle,
    ProveedorMaterialPrecio,
    ReporteKPIProduccion,
    RegistroScrapDefecto,
    RegistroUsoRecursoProduccion,
)

SHIFT_HOURS_PER_DAY = Decimal('8')
MACHINE_HOURLY_RATE = Decimal('180.00')
LABOR_HOURLY_RATE = Decimal('95.00')
ZERO = Decimal('0')
HUNDRED = Decimal('100')

STATUS_RULES = {
    'oee': {'direction': 'high', 'green': Decimal('85'), 'yellow': Decimal('70')},
    'tiempo_ciclo': {'direction': 'low_ratio', 'green': Decimal('1.10'), 'yellow': Decimal('1.25')},
    'tasa_rechazo': {'direction': 'low', 'green': Decimal('2'), 'yellow': Decimal('5')},
    'cumplimiento_ordenes': {'direction': 'high', 'green': Decimal('95'), 'yellow': Decimal('85')},
    'variacion_costos_pct': {'direction': 'low_abs', 'green': Decimal('5'), 'yellow': Decimal('10')},
    'utilizacion_recursos': {'direction': 'band', 'green_min': Decimal('70'), 'green_max': Decimal('90'), 'yellow_min': Decimal('55'), 'yellow_max': Decimal('95')},
}


def _to_decimal(value) -> Decimal:
    if isinstance(value, Decimal):
        return value
    if value in (None, ''):
        return ZERO
    return Decimal(str(value))


def _round(value: Decimal, places: str = '0.01') -> Decimal:
    return _to_decimal(value).quantize(Decimal(places), rounding=ROUND_HALF_UP)


def _safe_div(numerator: Decimal, denominator: Decimal) -> Decimal:
    denominator = _to_decimal(denominator)
    if denominator <= 0:
        return ZERO
    return _to_decimal(numerator) / denominator


def _hours_between(start: datetime | None, end: datetime | None) -> Decimal:
    if not start or not end or end <= start:
        return ZERO
    seconds = Decimal(str((end - start).total_seconds()))
    return seconds / Decimal('3600')


def _minutes_for_operation(operation: BOMOperacion) -> Decimal:
    tiempo = _to_decimal(operation.tiempo_estimado)
    if tiempo <= 0:
        return ZERO
    if operation.unidad_tiempo == BOMOperacion.UnidadTiempo.HORAS:
        return tiempo * Decimal('60')
    if operation.unidad_tiempo == BOMOperacion.UnidadTiempo.SEGUNDOS:
        return tiempo / Decimal('60')
    return tiempo


def _ideal_cycle_minutes_per_unit(bom) -> Decimal:
    total = ZERO
    for operation in bom.operaciones.all():
        total += _minutes_for_operation(operation)
    return total


def _weighted_operator_factor(bom) -> Decimal:
    weighted_minutes = ZERO
    total_minutes = ZERO
    for operation in bom.operaciones.all():
        minutes = _minutes_for_operation(operation)
        operadores = _to_decimal(operation.operadores_requeridos or 1)
        weighted_minutes += minutes * operadores
        total_minutes += minutes
    if total_minutes <= 0:
        return Decimal('1')
    return weighted_minutes / total_minutes


def _scheduled_hours_for_order(order: OrdenFabricacion, ideal_cycle_minutes: Decimal) -> Decimal:
    if order.fecha_inicio_programada and order.fecha_fin_programada and order.fecha_fin_programada >= order.fecha_inicio_programada:
        days = (order.fecha_fin_programada - order.fecha_inicio_programada).days + 1
        return Decimal(days) * SHIFT_HOURS_PER_DAY

    ideal_hours = (ideal_cycle_minutes * _to_decimal(order.cantidad_planificada)) / Decimal('60')
    if ideal_hours > 0:
        return ideal_hours
    return _hours_between(order.fecha_inicio_real, order.fecha_fin_real)


def _latest_material_costs(material_ids: set[int]) -> dict[int, Decimal]:
    cost_map: dict[int, Decimal] = {}
    if not material_ids:
        return cost_map

    prices = (
        ProveedorMaterialPrecio.objects
        .filter(material_id__in=material_ids)
        .order_by('material_id', '-fecha_actualizacion')
    )
    for price in prices:
        if price.material_id not in cost_map:
            cost_map[price.material_id] = _to_decimal(price.precio_unitario)
    return cost_map


def _material_cost(detail: OrdenFabricacionDetalle, cost_map: dict[int, Decimal], field_name: str) -> Decimal:
    quantity = _to_decimal(getattr(detail, field_name, ZERO))
    unit_cost = cost_map.get(detail.material_id, ZERO)
    return quantity * unit_cost


def _get_period_bounds(fecha_inicio: date | None, fecha_fin: date | None) -> tuple[date, date]:
    end_date = fecha_fin or timezone.localdate()
    start_date = fecha_inicio or (end_date - timedelta(days=29))
    if start_date > end_date:
        start_date, end_date = end_date, start_date
    return start_date, end_date


def _base_orders_queryset(fecha_inicio: date, fecha_fin: date):
    filtros = (
        Q(fecha_creacion__date__range=(fecha_inicio, fecha_fin)) |
        Q(fecha_actualizacion__date__range=(fecha_inicio, fecha_fin)) |
        Q(fecha_fin_real__date__range=(fecha_inicio, fecha_fin))
    )
    return (
        OrdenFabricacion.objects
        .filter(filtros)
        .select_related('bom', 'plan')
        .prefetch_related(
            'bom__operaciones',
            Prefetch('detalles', queryset=OrdenFabricacionDetalle.objects.select_related('material')),
        )
        .distinct()
    )


def _base_lots_queryset(fecha_inicio: date, fecha_fin: date):
    filtros = (
        Q(fecha_captura__range=(fecha_inicio, fecha_fin)) |
        Q(fecha_actualizacion__date__range=(fecha_inicio, fecha_fin))
    )
    return (
        LoteProduccion.objects
        .filter(filtros)
        .select_related('bom', 'orden_fabricacion', 'cliente_destino')
        .distinct()
    )


def _base_resource_usage_queryset(fecha_inicio: date, fecha_fin: date):
    return (
        RegistroUsoRecursoProduccion.objects
        .filter(fecha_creacion__date__range=(fecha_inicio, fecha_fin))
        .select_related('orden', 'costo_maquina', 'costo_operador__operador')
    )


def _base_scrap_queryset(fecha_inicio: date, fecha_fin: date):
    return (
        RegistroScrapDefecto.objects
        .filter(fecha_creacion__date__range=(fecha_inicio, fecha_fin))
        .select_related('orden', 'lote', 'informe_qa')
    )


def _produced_units(orders, lots) -> Decimal:
    total_lots = sum((_to_decimal(lot.cantidad_producida) for lot in lots), ZERO)
    if total_lots > 0:
        return total_lots
    return sum((_to_decimal(order.cantidad_producida) for order in orders), ZERO)


def _machine_rate_reference() -> dict:
    exact = {}
    line_totals = {}
    line_counts = {}
    overall_total = ZERO
    overall_count = 0

    for cost in CostoHoraMaquina.objects.filter(activo=True):
        rate = _to_decimal(cost.costo_hora)
        exact[(cost.linea_produccion or '', cost.maquina_nombre or '')] = rate
        line_totals[cost.linea_produccion] = line_totals.get(cost.linea_produccion, ZERO) + rate
        line_counts[cost.linea_produccion] = line_counts.get(cost.linea_produccion, 0) + 1
        overall_total += rate
        overall_count += 1

    line_avg = {
        line: _safe_div(total, Decimal(line_counts[line]))
        for line, total in line_totals.items()
        if line_counts[line]
    }
    overall_avg = _safe_div(overall_total, Decimal(overall_count)) if overall_count else MACHINE_HOURLY_RATE
    return {
        'exact': exact,
        'line_avg': line_avg,
        'overall_avg': overall_avg if overall_avg > 0 else MACHINE_HOURLY_RATE,
    }


def _average_operator_rate() -> Decimal:
    rates = [
        _to_decimal(item.costo_hora_real)
        for item in CostoHoraOperador.objects.filter(activo=True)
        if _to_decimal(item.costo_hora_real) > 0
    ]
    if not rates:
        return LABOR_HOURLY_RATE
    return _safe_div(sum(rates, ZERO), Decimal(len(rates)))


def _planned_machine_rate(order: OrdenFabricacion, rate_reference: dict) -> Decimal:
    matches = []
    for operation in order.bom.operaciones.all():
        key = (operation.linea_produccion or order.linea_produccion or '', operation.recurso_maquina or '')
        if operation.recurso_maquina and key in rate_reference['exact']:
            matches.append(rate_reference['exact'][key])

    if matches:
        return _safe_div(sum(matches, ZERO), Decimal(len(matches)))

    line_rate = rate_reference['line_avg'].get(order.linea_produccion)
    if line_rate and line_rate > 0:
        return line_rate
    return rate_reference['overall_avg'] or MACHINE_HOURLY_RATE
    return (
        LoteProduccion.objects
        .filter(filtros)
        .select_related('bom', 'orden_fabricacion', 'cliente_destino')
        .distinct()
    )


def calcular_oee(orders, lots, scrap_records=None) -> dict:
    actual_hours = ZERO
    planned_hours = ZERO
    ideal_hours_output = ZERO

    for order in orders:
        ideal_cycle_minutes = _ideal_cycle_minutes_per_unit(order.bom)
        actual_hours += _hours_between(order.fecha_inicio_real, order.fecha_fin_real)
        planned_hours += _scheduled_hours_for_order(order, ideal_cycle_minutes)
        ideal_hours_output += (ideal_cycle_minutes * _to_decimal(order.cantidad_producida)) / Decimal('60')

    total_qty = _produced_units(orders, lots)
    rejected_qty = ZERO
    if scrap_records:
        rejected_qty = sum((_to_decimal(scrap.cantidad_defectos) for scrap in scrap_records), ZERO)
        good_qty = max(total_qty - rejected_qty, ZERO)
    else:
        good_qty = sum((_to_decimal(lot.cantidad_producida) for lot in lots if lot.estado == LoteProduccion.EstadoLote.VALIDADO), ZERO)

    disponibilidad = _safe_div(actual_hours, planned_hours) * HUNDRED
    rendimiento = _safe_div(ideal_hours_output, actual_hours) * HUNDRED
    calidad = _safe_div(good_qty, total_qty) * HUNDRED
    oee = (disponibilidad / HUNDRED) * (rendimiento / HUNDRED) * (calidad / HUNDRED) * HUNDRED

    return {
        'disponibilidad': _round(min(disponibilidad, HUNDRED)),
        'rendimiento': _round(min(rendimiento, HUNDRED)),
        'calidad': _round(min(calidad, HUNDRED)),
        'oee': _round(min(oee, HUNDRED)),
        'planned_hours': _round(planned_hours),
        'actual_hours': _round(actual_hours),
        'good_qty': _round(good_qty),
        'total_qty': _round(total_qty),
        'rejected_qty': _round(rejected_qty),
    }


def calcular_tiempo_ciclo(orders) -> dict:
    runtime_hours = sum((_hours_between(order.fecha_inicio_real, order.fecha_fin_real) for order in orders), ZERO)
    total_units = sum((_to_decimal(order.cantidad_producida) for order in orders), ZERO)

    cycle_minutes = _safe_div(runtime_hours * Decimal('60'), total_units)

    ideal_cycle_acc = ZERO
    cycle_count = 0
    for order in orders:
        ideal_cycle = _ideal_cycle_minutes_per_unit(order.bom)
        if ideal_cycle > 0:
            ideal_cycle_acc += ideal_cycle
            cycle_count += 1
    ideal_cycle_avg = _safe_div(ideal_cycle_acc, Decimal(cycle_count)) if cycle_count else ZERO

    return {
        'tiempo_ciclo_promedio': _round(cycle_minutes),
        'tiempo_ciclo_ideal': _round(ideal_cycle_avg),
    }


def calcular_tasa_rechazo(orders, lots, scrap_records=None) -> dict:
    total_qty = _produced_units(orders, lots)
    if scrap_records:
        rejected_qty = sum((_to_decimal(scrap.cantidad_defectos) for scrap in scrap_records), ZERO)
    else:
        rejected_qty = sum((_to_decimal(lot.cantidad_producida) for lot in lots if lot.estado == LoteProduccion.EstadoLote.RECHAZADO), ZERO)
    rate = _safe_div(rejected_qty, total_qty) * HUNDRED
    return {
        'tasa_rechazo': _round(rate),
        'unidades_rechazadas': _round(rejected_qty),
        'unidades_producidas': _round(total_qty),
    }


def calcular_cumplimiento_ordenes(orders) -> dict:
    completed = [order for order in orders if order.estado == OrdenFabricacion.EstadoOF.COMPLETADA]
    planned = len(completed)
    on_time = 0
    for order in completed:
        if order.fecha_fin_programada and order.fecha_fin_real and order.fecha_fin_real.date() <= order.fecha_fin_programada:
            on_time += 1
    rate = _safe_div(Decimal(on_time), Decimal(planned)) * HUNDRED if planned else ZERO
    return {
        'cumplimiento_ordenes': _round(rate),
        'ordenes_a_tiempo': on_time,
        'ordenes_completadas': planned,
    }


def calcular_costos(orders, resource_usage_by_order=None) -> dict:
    material_ids = set()
    for order in orders:
        for detail in order.detalles.all():
            material_ids.add(detail.material_id)
    cost_map = _latest_material_costs(material_ids)
    machine_rate_reference = _machine_rate_reference()
    average_operator_rate = _average_operator_rate()

    planned_cost = ZERO
    real_cost = ZERO
    planned_material_cost = ZERO
    real_material_cost = ZERO
    planned_resource_cost = ZERO
    real_resource_cost = ZERO

    for order in orders:
        ideal_cycle_minutes = _ideal_cycle_minutes_per_unit(order.bom)
        weighted_operators = _weighted_operator_factor(order.bom)
        planned_hours = _scheduled_hours_for_order(order, ideal_cycle_minutes)
        actual_hours = _hours_between(order.fecha_inicio_real, order.fecha_fin_real)
        planned_machine_rate = _planned_machine_rate(order, machine_rate_reference)

        for detail in order.detalles.all():
            planned_detail_cost = _material_cost(detail, cost_map, 'cantidad_requerida')
            real_detail_cost = _material_cost(detail, cost_map, 'cantidad_consumida')
            planned_material_cost += planned_detail_cost
            real_material_cost += real_detail_cost

        planned_machine_cost = planned_hours * planned_machine_rate
        planned_labor_cost = planned_hours * weighted_operators * average_operator_rate
        planned_resource_cost += planned_machine_cost + planned_labor_cost

        usages = (resource_usage_by_order or {}).get(order.id, [])
        machine_usages = [usage for usage in usages if usage.tipo_recurso == RegistroUsoRecursoProduccion.TipoRecurso.MAQUINA]
        operator_usages = [usage for usage in usages if usage.tipo_recurso == RegistroUsoRecursoProduccion.TipoRecurso.OPERADOR]

        real_machine_cost = sum((_to_decimal(usage.costo_total) for usage in machine_usages), ZERO)
        real_labor_cost = sum((_to_decimal(usage.costo_total) for usage in operator_usages), ZERO)
        if not machine_usages:
            real_machine_cost = actual_hours * planned_machine_rate
        if not operator_usages:
            real_labor_cost = actual_hours * weighted_operators * average_operator_rate

        real_resource_cost += real_machine_cost + real_labor_cost

    planned_cost = planned_material_cost + planned_resource_cost
    real_cost = real_material_cost + real_resource_cost

    variation = real_cost - planned_cost
    variation_pct = _safe_div(variation, planned_cost) * HUNDRED if planned_cost > 0 else ZERO

    return {
        'costo_planificado': _round(planned_cost),
        'costo_real': _round(real_cost),
        'variacion_costos': _round(variation),
        'variacion_costos_pct': _round(variation_pct),
        'costo_planificado_materiales': _round(planned_material_cost),
        'costo_real_materiales': _round(real_material_cost),
        'costo_planificado_recursos': _round(planned_resource_cost),
        'costo_real_recursos': _round(real_resource_cost),
    }


def calcular_utilizacion_recursos(orders, resource_usage_by_order=None) -> dict:
    planned_machine_hours = ZERO
    actual_machine_hours = ZERO
    planned_labor_hours = ZERO
    actual_labor_hours = ZERO

    for order in orders:
        ideal_cycle_minutes = _ideal_cycle_minutes_per_unit(order.bom)
        weighted_operators = _weighted_operator_factor(order.bom)
        planned_hours = _scheduled_hours_for_order(order, ideal_cycle_minutes)
        actual_hours = _hours_between(order.fecha_inicio_real, order.fecha_fin_real)
        usages = (resource_usage_by_order or {}).get(order.id, [])
        machine_hours = sum((_to_decimal(usage.horas_reales) for usage in usages if usage.tipo_recurso == RegistroUsoRecursoProduccion.TipoRecurso.MAQUINA), ZERO)
        operator_hours = sum((_to_decimal(usage.horas_reales) for usage in usages if usage.tipo_recurso == RegistroUsoRecursoProduccion.TipoRecurso.OPERADOR), ZERO)

        planned_machine_hours += planned_hours
        actual_machine_hours += machine_hours if machine_hours > 0 else actual_hours
        planned_labor_hours += planned_hours * weighted_operators
        actual_labor_hours += operator_hours if operator_hours > 0 else actual_hours * weighted_operators

    machine_util = _safe_div(actual_machine_hours, planned_machine_hours) * HUNDRED
    labor_util = _safe_div(actual_labor_hours, planned_labor_hours) * HUNDRED
    overall_util = (machine_util + labor_util) / Decimal('2') if planned_machine_hours > 0 or planned_labor_hours > 0 else ZERO

    return {
        'utilizacion_maquinas': _round(min(machine_util, HUNDRED)),
        'utilizacion_personal': _round(min(labor_util, HUNDRED)),
        'utilizacion_recursos': _round(min(overall_util, HUNDRED)),
    }


def estado_kpi(codigo: str, valor: Decimal, referencia: Decimal | None = None) -> str:
    rules = STATUS_RULES[codigo]
    value = _to_decimal(valor)

    if rules['direction'] == 'high':
        if value >= rules['green']:
            return 'verde'
        if value >= rules['yellow']:
            return 'amarillo'
        return 'rojo'

    if rules['direction'] == 'low':
        if value <= rules['green']:
            return 'verde'
        if value <= rules['yellow']:
            return 'amarillo'
        return 'rojo'

    if rules['direction'] == 'low_abs':
        abs_value = abs(value)
        if abs_value <= rules['green']:
            return 'verde'
        if abs_value <= rules['yellow']:
            return 'amarillo'
        return 'rojo'

    if rules['direction'] == 'low_ratio':
        ratio = _safe_div(value, referencia or Decimal('1')) if referencia else Decimal('999')
        if ratio <= rules['green']:
            return 'verde'
        if ratio <= rules['yellow']:
            return 'amarillo'
        return 'rojo'

    if rules['direction'] == 'band':
        if rules['green_min'] <= value <= rules['green_max']:
            return 'verde'
        if rules['yellow_min'] <= value <= rules['yellow_max']:
            return 'amarillo'
        return 'rojo'

    return 'amarillo'


def calcular_kpis_produccion(fecha_inicio: date | None = None, fecha_fin: date | None = None) -> dict:
    start_date, end_date = _get_period_bounds(fecha_inicio, fecha_fin)
    orders = list(_base_orders_queryset(start_date, end_date))
    lots = list(_base_lots_queryset(start_date, end_date))
    order_ids = [order.id for order in orders]
    lot_ids = [lot.id for lot in lots]

    resource_usages = list(_base_resource_usage_queryset(start_date, end_date).filter(orden_id__in=order_ids)) if order_ids else []
    resource_usage_by_order = {}
    for usage in resource_usages:
        resource_usage_by_order.setdefault(usage.orden_id, []).append(usage)

    scrap_queryset = _base_scrap_queryset(start_date, end_date)
    scrap_filters = Q()
    if order_ids:
        scrap_filters |= Q(orden_id__in=order_ids)
    if lot_ids:
        scrap_filters |= Q(lote_id__in=lot_ids)
    scrap_records = list(scrap_queryset.filter(scrap_filters)) if scrap_filters else []

    qa_validated = 0
    qa_pending = 0
    machine_failure_qty = ZERO
    for scrap in scrap_records:
        if hasattr(scrap, 'informe_qa'):
            if scrap.informe_qa.resultado_validacion == InformeValidacionDefectoQA.ResultadoValidacion.VALIDADO:
                qa_validated += 1
            if scrap.informe_qa.falla_maquina:
                machine_failure_qty += _to_decimal(scrap.cantidad_defectos)
        else:
            qa_pending += 1

    oee_data = calcular_oee(orders, lots, scrap_records=scrap_records)
    cycle_data = calcular_tiempo_ciclo(orders)
    reject_data = calcular_tasa_rechazo(orders, lots, scrap_records=scrap_records)
    compliance_data = calcular_cumplimiento_ordenes(orders)
    cost_data = calcular_costos(orders, resource_usage_by_order=resource_usage_by_order)
    resource_data = calcular_utilizacion_recursos(orders, resource_usage_by_order=resource_usage_by_order)

    metrics = {
        'oee': {
            'label': 'OEE',
            'value': oee_data['oee'],
            'unit': '%',
            'status': estado_kpi('oee', oee_data['oee']),
        },
        'tiempo_ciclo': {
            'label': 'Tiempo de ciclo',
            'value': cycle_data['tiempo_ciclo_promedio'],
            'unit': 'min/ud',
            'status': estado_kpi('tiempo_ciclo', cycle_data['tiempo_ciclo_promedio'], cycle_data['tiempo_ciclo_ideal']),
        },
        'tasa_rechazo': {
            'label': 'Tasa de rechazo',
            'value': reject_data['tasa_rechazo'],
            'unit': '%',
            'status': estado_kpi('tasa_rechazo', reject_data['tasa_rechazo']),
        },
        'cumplimiento_ordenes': {
            'label': 'Cumplimiento de órdenes',
            'value': compliance_data['cumplimiento_ordenes'],
            'unit': '%',
            'status': estado_kpi('cumplimiento_ordenes', compliance_data['cumplimiento_ordenes']),
        },
        'variacion_costos_pct': {
            'label': 'Variación de costos',
            'value': cost_data['variacion_costos_pct'],
            'unit': '%',
            'status': estado_kpi('variacion_costos_pct', cost_data['variacion_costos_pct']),
        },
        'utilizacion_recursos': {
            'label': 'Utilización de recursos',
            'value': resource_data['utilizacion_recursos'],
            'unit': '%',
            'status': estado_kpi('utilizacion_recursos', resource_data['utilizacion_recursos']),
        },
    }

    alerts = []
    for code, metric in metrics.items():
        if metric['status'] == 'rojo':
            alerts.append(f"{metric['label']} en nivel crítico: {metric['value']}{metric['unit']}")
    if qa_pending > 0:
        alerts.append(f'Defectos pendientes de validación QA: {qa_pending}')
    if machine_failure_qty > 0:
        alerts.append(f'Defectos validados por falla de máquina: {_round(machine_failure_qty)} piezas')

    detail = {
        'total_ordenes': len(orders),
        'ordenes_completadas': compliance_data['ordenes_completadas'],
        'ordenes_a_tiempo': compliance_data['ordenes_a_tiempo'],
        'total_lotes': len(lots),
        'unidades_buenas': float(oee_data['good_qty']),
        'unidades_totales': float(oee_data['total_qty']),
        'unidades_rechazadas': float(reject_data['unidades_rechazadas']),
        'defectos_registrados': float(sum((_to_decimal(scrap.cantidad_defectos) for scrap in scrap_records), ZERO)),
        'defectos_validados_qa': qa_validated,
        'defectos_pendientes_qa': qa_pending,
        'defectos_falla_maquina': float(machine_failure_qty),
        'horas_programadas': float(oee_data['planned_hours']),
        'horas_reales': float(oee_data['actual_hours']),
        'tiempo_ciclo_ideal': float(cycle_data['tiempo_ciclo_ideal']),
        'costo_variacion_pct': float(cost_data['variacion_costos_pct']),
        'costo_planificado_materiales': float(cost_data['costo_planificado_materiales']),
        'costo_real_materiales': float(cost_data['costo_real_materiales']),
        'costo_planificado_recursos': float(cost_data['costo_planificado_recursos']),
        'costo_real_recursos': float(cost_data['costo_real_recursos']),
        'ordenes_recientes': [
            {
                'folio': order.folio,
                'producto': order.bom.producto,
                'estado': order.get_estado_display(),
                'cantidad_planificada': float(_to_decimal(order.cantidad_planificada)),
                'cantidad_producida': float(_to_decimal(order.cantidad_producida)),
            }
            for order in orders[:8]
        ],
        'scrap_reciente': [
            {
                'referencia': scrap.orden.folio if scrap.orden_id else (scrap.lote.folio if scrap.lote_id else 'Sin referencia'),
                'tipo_defecto': scrap.get_tipo_defecto_display(),
                'cantidad': float(_to_decimal(scrap.cantidad_defectos)),
                'causa': scrap.causa,
                'qa': scrap.informe_qa.get_resultado_validacion_display() if hasattr(scrap, 'informe_qa') else 'Pendiente QA',
                'falla_maquina': bool(scrap.informe_qa.falla_maquina) if hasattr(scrap, 'informe_qa') else False,
            }
            for scrap in scrap_records[:10]
        ],
    }

    return {
        'fecha_inicio': start_date,
        'fecha_fin': end_date,
        'metrics': metrics,
        'alerts': alerts,
        'oee': oee_data,
        'cycle': cycle_data,
        'reject': reject_data,
        'compliance': compliance_data,
        'cost': cost_data,
        'resources': resource_data,
        'detail': detail,
    }


def generar_reporte_kpis_produccion(usuario=None, fecha_inicio: date | None = None, fecha_fin: date | None = None) -> ReporteKPIProduccion:
    data = calcular_kpis_produccion(fecha_inicio=fecha_inicio, fecha_fin=fecha_fin)
    report = ReporteKPIProduccion.objects.create(
        fecha_inicio=data['fecha_inicio'],
        fecha_fin=data['fecha_fin'],
        disponibilidad=data['oee']['disponibilidad'],
        rendimiento=data['oee']['rendimiento'],
        calidad=data['oee']['calidad'],
        oee=data['oee']['oee'],
        tiempo_ciclo_promedio=data['cycle']['tiempo_ciclo_promedio'],
        tasa_rechazo=data['reject']['tasa_rechazo'],
        cumplimiento_ordenes=data['compliance']['cumplimiento_ordenes'],
        costo_planificado=data['cost']['costo_planificado'],
        costo_real=data['cost']['costo_real'],
        variacion_costos=data['cost']['variacion_costos'],
        utilizacion_maquinas=data['resources']['utilizacion_maquinas'],
        utilizacion_personal=data['resources']['utilizacion_personal'],
        utilizacion_recursos=data['resources']['utilizacion_recursos'],
        alertas=data['alerts'],
        detalle=data['detail'],
        generado_por=usuario,
    )
    return report


def exportar_reporte_excel(reporte: ReporteKPIProduccion) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'KPIs MFG'

    sheet['A1'] = 'Reporte KPIs de Producción MFG'
    sheet['A1'].font = Font(bold=True, size=14)
    sheet['A2'] = f"Periodo: {reporte.fecha_inicio} a {reporte.fecha_fin}"

    rows = [
        ('Disponibilidad', float(reporte.disponibilidad), '%'),
        ('Rendimiento', float(reporte.rendimiento), '%'),
        ('Calidad', float(reporte.calidad), '%'),
        ('OEE', float(reporte.oee), '%'),
        ('Tiempo de ciclo promedio', float(reporte.tiempo_ciclo_promedio), 'min/ud'),
        ('Tasa de rechazo', float(reporte.tasa_rechazo), '%'),
        ('Cumplimiento de órdenes', float(reporte.cumplimiento_ordenes), '%'),
        ('Costo planificado', float(reporte.costo_planificado), 'MXN'),
        ('Costo real', float(reporte.costo_real), 'MXN'),
        ('Variación de costos', float(reporte.variacion_costos), 'MXN'),
        ('Utilización de máquinas', float(reporte.utilizacion_maquinas), '%'),
        ('Utilización de personal', float(reporte.utilizacion_personal), '%'),
        ('Utilización de recursos', float(reporte.utilizacion_recursos), '%'),
    ]

    sheet.append([])
    sheet.append(['Indicador', 'Valor', 'Unidad'])
    for cell in sheet[4]:
        cell.font = Font(bold=True)
    for row in rows:
        sheet.append(list(row))

    sheet.append([])
    sheet.append(['Alertas'])
    sheet[sheet.max_row][0].font = Font(bold=True)
    for alert in reporte.alertas or ['Sin alertas críticas']:
        sheet.append([alert])

    detail = reporte.detalle or {}
    sheet.append([])
    sheet.append(['Scrap y QA', 'Valor'])
    for cell in sheet[sheet.max_row]:
        cell.font = Font(bold=True)
    sheet.append(['Defectos registrados', detail.get('defectos_registrados', 0)])
    sheet.append(['Defectos validados QA', detail.get('defectos_validados_qa', 0)])
    sheet.append(['Defectos pendientes QA', detail.get('defectos_pendientes_qa', 0)])
    sheet.append(['Defectos por falla de máquina', detail.get('defectos_falla_maquina', 0)])

    sheet.append([])
    sheet.append(['Desglose de costos', 'Valor'])
    for cell in sheet[sheet.max_row]:
        cell.font = Font(bold=True)
    sheet.append(['Costo planificado materiales', detail.get('costo_planificado_materiales', 0)])
    sheet.append(['Costo real materiales', detail.get('costo_real_materiales', 0)])
    sheet.append(['Costo planificado recursos', detail.get('costo_planificado_recursos', 0)])
    sheet.append(['Costo real recursos', detail.get('costo_real_recursos', 0)])

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def exportar_reporte_pdf(reporte: ReporteKPIProduccion) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    output = BytesIO()
    document = SimpleDocTemplate(output, pagesize=A4)
    styles = getSampleStyleSheet()
    story = [
        Paragraph('Reporte KPIs de Producción MFG', styles['Title']),
        Paragraph(f'Periodo: {reporte.fecha_inicio} a {reporte.fecha_fin}', styles['BodyText']),
        Spacer(1, 12),
    ]

    table_data = [
        ['Indicador', 'Valor', 'Unidad'],
        ['Disponibilidad', f'{reporte.disponibilidad}%', '%'],
        ['Rendimiento', f'{reporte.rendimiento}%', '%'],
        ['Calidad', f'{reporte.calidad}%', '%'],
        ['OEE', f'{reporte.oee}%', '%'],
        ['Tiempo de ciclo promedio', f'{reporte.tiempo_ciclo_promedio}', 'min/ud'],
        ['Tasa de rechazo', f'{reporte.tasa_rechazo}%', '%'],
        ['Cumplimiento de órdenes', f'{reporte.cumplimiento_ordenes}%', '%'],
        ['Costo planificado', f'${reporte.costo_planificado}', 'MXN'],
        ['Costo real', f'${reporte.costo_real}', 'MXN'],
        ['Variación de costos', f'${reporte.variacion_costos}', 'MXN'],
        ['Utilización de máquinas', f'{reporte.utilizacion_maquinas}%', '%'],
        ['Utilización de personal', f'{reporte.utilizacion_personal}%', '%'],
        ['Utilización de recursos', f'{reporte.utilizacion_recursos}%', '%'],
    ]
    table = Table(table_data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f3e67')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#dce7f5')),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f8fbff')),
    ]))
    story.append(table)
    story.append(Spacer(1, 12))
    story.append(Paragraph('Alertas automáticas', styles['Heading2']))
    for alert in reporte.alertas or ['Sin alertas críticas']:
        story.append(Paragraph(f'- {alert}', styles['BodyText']))

    detail = reporte.detalle or {}
    story.append(Spacer(1, 12))
    story.append(Paragraph('Resumen Scrap y QA', styles['Heading2']))
    qa_table = Table([
        ['Concepto', 'Valor'],
        ['Defectos registrados', detail.get('defectos_registrados', 0)],
        ['Defectos validados QA', detail.get('defectos_validados_qa', 0)],
        ['Defectos pendientes QA', detail.get('defectos_pendientes_qa', 0)],
        ['Defectos por falla de máquina', detail.get('defectos_falla_maquina', 0)],
    ], repeatRows=1)
    qa_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3f6b96')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#dce7f5')),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f8fbff')),
    ]))
    story.append(qa_table)
    story.append(Spacer(1, 12))
    story.append(Paragraph('Desglose de costos', styles['Heading2']))
    cost_table = Table([
        ['Concepto', 'Valor MXN'],
        ['Costo planificado materiales', detail.get('costo_planificado_materiales', 0)],
        ['Costo real materiales', detail.get('costo_real_materiales', 0)],
        ['Costo planificado recursos', detail.get('costo_planificado_recursos', 0)],
        ['Costo real recursos', detail.get('costo_real_recursos', 0)],
    ], repeatRows=1)
    cost_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3f6b96')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#dce7f5')),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f8fbff')),
    ]))
    story.append(cost_table)

    document.build(story)
    return output.getvalue()
