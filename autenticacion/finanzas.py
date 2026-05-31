from __future__ import annotations

from datetime import date, timedelta
from decimal import Decimal, ROUND_HALF_UP
from io import BytesIO

from django.db.models import Prefetch, Q, Sum
from django.utils import timezone

from .kpi_produccion import calcular_kpis_produccion
from .models import (
    CosteoProduccion,
    CuentaPorPagarCobrar,
    DeclaracionImpuesto,
    EstadoFinanciero,
    LoteProduccion,
    OrdenCompra,
    OrdenFabricacion,
    OrdenFabricacionDetalle,
    PresupuestoFinanciero,
    ProveedorMaterialPrecio,
    RegistroUsoRecursoProduccion,
    ReporteFinanciero,
)

ZERO = Decimal('0')
HUNDRED = Decimal('100')


def _to_decimal(value) -> Decimal:
    if isinstance(value, Decimal):
        return value
    if value in (None, ''):
        return ZERO
    return Decimal(str(value))


def _round(value: Decimal, places: str = '0.01') -> Decimal:
    return _to_decimal(value).quantize(Decimal(places), rounding=ROUND_HALF_UP)


def _safe_div(numerator: Decimal, denominator: Decimal) -> Decimal:
    denominator = _to_decimal(denominator)
    if denominator <= 0:
        return ZERO
    return _to_decimal(numerator) / denominator


def _get_period_bounds(fecha_inicio: date | None = None, fecha_fin: date | None = None) -> tuple[date, date]:
    end_date = fecha_fin or timezone.localdate()
    start_date = fecha_inicio or (end_date - timedelta(days=29))
    if start_date > end_date:
        start_date, end_date = end_date, start_date
    return start_date, end_date


def _latest_material_costs(material_ids: set[int]) -> dict[int, Decimal]:
    cost_map: dict[int, Decimal] = {}
    if not material_ids:
        return cost_map

    precios = (
        ProveedorMaterialPrecio.objects
        .filter(material_id__in=material_ids)
        .order_by('material_id', '-fecha_actualizacion')
    )
    for precio in precios:
        if precio.material_id not in cost_map:
            cost_map[precio.material_id] = _to_decimal(precio.precio_unitario)
    return cost_map


def _orders_in_period(fecha_inicio: date, fecha_fin: date):
    filtros = (
        Q(fecha_creacion__date__range=(fecha_inicio, fecha_fin)) |
        Q(fecha_actualizacion__date__range=(fecha_inicio, fecha_fin)) |
        Q(fecha_fin_real__date__range=(fecha_inicio, fecha_fin))
    )
    return (
        OrdenFabricacion.objects
        .filter(filtros)
        .select_related('bom', 'plan')
        .prefetch_related(
            Prefetch('detalles', queryset=OrdenFabricacionDetalle.objects.select_related('material')),
            'lotes_produccion',
            'usos_recursos',
        )
        .distinct()
    )


def _lots_in_period(fecha_inicio: date, fecha_fin: date):
    return (
        LoteProduccion.objects
        .filter(
            Q(fecha_captura__range=(fecha_inicio, fecha_fin)) |
            Q(fecha_actualizacion__date__range=(fecha_inicio, fecha_fin))
        )
        .select_related('bom', 'orden_fabricacion')
        .distinct()
    )


def consolidar_costeos_produccion(usuario=None, fecha_inicio: date | None = None, fecha_fin: date | None = None) -> list[CosteoProduccion]:
    start_date, end_date = _get_period_bounds(fecha_inicio, fecha_fin)
    orders = list(_orders_in_period(start_date, end_date))
    lotes = list(_lots_in_period(start_date, end_date))
    lotes_por_orden = {}
    for lote in lotes:
        if lote.orden_fabricacion_id:
            lotes_por_orden.setdefault(lote.orden_fabricacion_id, []).append(lote)

    material_ids = {
        detalle.material_id
        for order in orders
        for detalle in order.detalles.all()
    }
    material_costs = _latest_material_costs(material_ids)
    registros = []

    for order in orders:
        costo_material_plan = ZERO
        costo_material_real = ZERO
        for detalle in order.detalles.all():
            costo_unitario = material_costs.get(detalle.material_id, ZERO)
            costo_material_plan += _to_decimal(detalle.cantidad_requerida) * costo_unitario
            costo_material_real += _to_decimal(detalle.cantidad_consumida) * costo_unitario

        usos = list(order.usos_recursos.all())
        costo_maquina_real = sum(
            (_to_decimal(uso.costo_total) for uso in usos if uso.tipo_recurso == RegistroUsoRecursoProduccion.TipoRecurso.MAQUINA),
            ZERO,
        )
        costo_operador_real = sum(
            (_to_decimal(uso.costo_total) for uso in usos if uso.tipo_recurso == RegistroUsoRecursoProduccion.TipoRecurso.OPERADOR),
            ZERO,
        )
        costo_total_plan = costo_material_plan
        costo_total_real = costo_material_real + costo_maquina_real + costo_operador_real
        ingreso_estimado = _to_decimal(order.cantidad_producida) * Decimal('12.50')
        rentabilidad = ingreso_estimado - costo_total_real
        margen_pct = _safe_div(rentabilidad, ingreso_estimado) * HUNDRED if ingreso_estimado > 0 else ZERO

        defaults = {
            'lote_produccion': None,
            'costo_material_plan': _round(costo_material_plan),
            'costo_material_real': _round(costo_material_real),
            'costo_maquina_real': _round(costo_maquina_real),
            'costo_operador_real': _round(costo_operador_real),
            'costo_total_plan': _round(costo_total_plan),
            'costo_total_real': _round(costo_total_real),
            'ingreso_estimado': _round(ingreso_estimado),
            'rentabilidad': _round(rentabilidad),
            'margen_pct': _round(margen_pct),
            'estado': CosteoProduccion.EstadoCosteo.CERRADO,
            'detalle': {
                'producto': order.bom.producto,
                'cantidad_planificada': float(_to_decimal(order.cantidad_planificada)),
                'cantidad_producida': float(_to_decimal(order.cantidad_producida)),
                'lotes_relacionados': [lote.folio for lote in lotes_por_orden.get(order.id, [])],
            },
            'creado_por': usuario or order.creado_por,
            'actualizado_por': usuario or order.creado_por,
        }
        costeo, _ = CosteoProduccion.objects.update_or_create(
            orden_fabricacion=order,
            lote_produccion=None,
            defaults=defaults,
        )
        registros.append(costeo)

        lotes_orden = lotes_por_orden.get(order.id, [])
        total_lote_qty = sum((_to_decimal(lote.cantidad_producida) for lote in lotes_orden), ZERO)
        for lote in lotes_orden:
            proporcion = _safe_div(_to_decimal(lote.cantidad_producida), total_lote_qty) if total_lote_qty > 0 else ZERO
            lote_defaults = {
                'orden_fabricacion': order,
                'costo_material_plan': _round(costo_material_plan * proporcion),
                'costo_material_real': _round(costo_material_real * proporcion),
                'costo_maquina_real': _round(costo_maquina_real * proporcion),
                'costo_operador_real': _round(costo_operador_real * proporcion),
                'costo_total_plan': _round(costo_total_plan * proporcion),
                'costo_total_real': _round(costo_total_real * proporcion),
                'ingreso_estimado': _round(ingreso_estimado * proporcion),
                'rentabilidad': _round(rentabilidad * proporcion),
                'margen_pct': _round(margen_pct),
                'estado': CosteoProduccion.EstadoCosteo.CERRADO,
                'detalle': {
                    'producto': lote.bom.producto,
                    'cantidad_producida': float(_to_decimal(lote.cantidad_producida)),
                    'cliente_destino': lote.cliente_destino.nombre if lote.cliente_destino_id else '',
                },
                'creado_por': usuario or lote.creado_por,
                'actualizado_por': usuario or lote.creado_por,
            }
            lote_costeo, _ = CosteoProduccion.objects.update_or_create(
                lote_produccion=lote,
                defaults=lote_defaults,
            )
            registros.append(lote_costeo)

    return registros


def calcular_dashboard_finanzas(fecha_inicio: date | None = None, fecha_fin: date | None = None) -> dict:
    start_date, end_date = _get_period_bounds(fecha_inicio, fecha_fin)
    cuentas = list(
        CuentaPorPagarCobrar.objects.filter(fecha_emision__range=(start_date, end_date)).select_related('orden_compra', 'proveedor', 'cliente_compra')
    )
    presupuestos = list(PresupuestoFinanciero.objects.filter(fecha_inicio__lte=end_date, fecha_fin__gte=start_date, activo=True))
    ordenes_compra = list(OrdenCompra.objects.filter(fecha_orden__range=(start_date, end_date)).select_related('proveedor'))
    kpis_mfg = calcular_kpis_produccion(fecha_inicio=start_date, fecha_fin=end_date)
    costeos = list(CosteoProduccion.objects.filter(fecha_actualizacion__date__range=(start_date, end_date), lote_produccion__isnull=True).select_related('orden_fabricacion', 'orden_fabricacion__bom'))
    declaraciones = list(DeclaracionImpuesto.objects.filter(periodo_inicio__lte=end_date, periodo_fin__gte=start_date))

    ingresos = sum((_to_decimal(c.monto_pagado) for c in cuentas if c.tipo == CuentaPorPagarCobrar.TipoCuenta.POR_COBRAR), ZERO)
    egresos = sum((_to_decimal(c.monto_pagado) for c in cuentas if c.tipo == CuentaPorPagarCobrar.TipoCuenta.POR_PAGAR), ZERO)
    flujo_caja = ingresos - egresos

    presupuesto_total = sum((_to_decimal(p.monto_presupuestado) for p in presupuestos), ZERO)
    gasto_real = sum((_to_decimal(p.monto_real) for p in presupuestos if p.categoria in {PresupuestoFinanciero.Categoria.GASTO, PresupuestoFinanciero.Categoria.OPEX, PresupuestoFinanciero.Categoria.CAPEX}), ZERO)
    ingreso_presupuestado = sum((_to_decimal(p.monto_presupuestado) for p in presupuestos if p.categoria == PresupuestoFinanciero.Categoria.INGRESO), ZERO)
    rentabilidad_total = sum((_to_decimal(c.rentabilidad) for c in costeos), ZERO)
    ingreso_estimado_total = sum((_to_decimal(c.ingreso_estimado) for c in costeos), ZERO)
    rentabilidad_pct = _safe_div(rentabilidad_total, ingreso_estimado_total) * HUNDRED if ingreso_estimado_total > 0 else ZERO

    oc_pendientes = sum(1 for oc in ordenes_compra if oc.estado in {OrdenCompra.EstadoOrden.BORRADOR, OrdenCompra.EstadoOrden.APROBADA, OrdenCompra.EstadoOrden.ENVIADA, OrdenCompra.EstadoOrden.PARCIAL})
    oc_completadas = sum(1 for oc in ordenes_compra if oc.estado == OrdenCompra.EstadoOrden.RECIBIDA)
    impuestos_pendientes = sum(1 for item in declaraciones if item.estado != DeclaracionImpuesto.EstadoDeclaracion.PRESENTADA)

    alerts = []
    if flujo_caja < 0:
        alerts.append('Flujo de caja en negativo en el periodo seleccionado.')
    if presupuesto_total > 0 and gasto_real > presupuesto_total:
        alerts.append('El gasto real supera el presupuesto total activo.')
    if oc_pendientes > oc_completadas and oc_pendientes > 0:
        alerts.append('Hay más órdenes de compra pendientes que completadas.')
    if impuestos_pendientes > 0:
        alerts.append(f'Existen {impuestos_pendientes} declaraciones fiscales pendientes.')

    return {
        'fecha_inicio': start_date,
        'fecha_fin': end_date,
        'kpis': {
            'flujo_caja': _round(flujo_caja),
            'rentabilidad': _round(rentabilidad_total),
            'rentabilidad_pct': _round(rentabilidad_pct),
            'presupuesto_total': _round(presupuesto_total),
            'gasto_real': _round(gasto_real),
            'ingreso_presupuestado': _round(ingreso_presupuestado),
            'oc_pendientes': oc_pendientes,
            'oc_completadas': oc_completadas,
            'costeos_generados': len(costeos),
            'impuestos_pendientes': impuestos_pendientes,
        },
        'produccion': {
            'costo_planificado': kpis_mfg['cost']['costo_planificado'],
            'costo_real': kpis_mfg['cost']['costo_real'],
            'variacion_costos': kpis_mfg['cost']['variacion_costos'],
            'oee': kpis_mfg['oee']['oee'],
            'tasa_rechazo': kpis_mfg['reject']['tasa_rechazo'],
        },
        'charts': {
            'presupuesto_labels': ['Presupuesto', 'Gasto real', 'Ingreso presupuestado'],
            'presupuesto_values': [float(_round(presupuesto_total)), float(_round(gasto_real)), float(_round(ingreso_presupuestado))],
            'costos_labels': ['Plan producción', 'Real producción', 'Rentabilidad'],
            'costos_values': [float(kpis_mfg['cost']['costo_planificado']), float(kpis_mfg['cost']['costo_real']), float(_round(rentabilidad_total))],
            'oc_labels': ['Pendientes', 'Completadas'],
            'oc_values': [oc_pendientes, oc_completadas],
        },
        'ordenes_compra_recientes': ordenes_compra[:8],
        'costeos_recientes': costeos[:8],
        'cuentas_recientes': cuentas[:8],
        'presupuestos': presupuestos[:8],
        'alertas': alerts,
    }


def generar_reporte_financiero(usuario=None, tipo: str = ReporteFinanciero.TipoReporte.KPI, fecha_inicio: date | None = None, fecha_fin: date | None = None) -> ReporteFinanciero:
    data = calcular_dashboard_finanzas(fecha_inicio=fecha_inicio, fecha_fin=fecha_fin)
    nombre = f'Reporte financiero {data["fecha_inicio"]} - {data["fecha_fin"]}'
    reporte = ReporteFinanciero.objects.create(
        nombre=nombre,
        tipo=tipo,
        fecha_inicio=data['fecha_inicio'],
        fecha_fin=data['fecha_fin'],
        indicadores={
            'kpis': {key: float(value) if isinstance(value, Decimal) else value for key, value in data['kpis'].items()},
            'produccion': {key: float(value) if isinstance(value, Decimal) else value for key, value in data['produccion'].items()},
            'charts': data['charts'],
        },
        alertas=data['alertas'],
        creado_por=usuario,
        actualizado_por=usuario,
    )
    return reporte


def exportar_reporte_financiero_excel(reporte: ReporteFinanciero) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    indicadores = reporte.indicadores or {}
    kpis = indicadores.get('kpis', {})
    produccion = indicadores.get('produccion', {})

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Finanzas'
    sheet['A1'] = 'Reporte Financiero'
    sheet['A1'].font = Font(bold=True, size=14)
    sheet['A2'] = f'Periodo: {reporte.fecha_inicio} a {reporte.fecha_fin}'
    sheet.append([])
    sheet.append(['KPI', 'Valor'])
    for cell in sheet[4]:
        cell.font = Font(bold=True)
    for key, value in kpis.items():
        sheet.append([key.replace('_', ' ').title(), value])

    sheet.append([])
    sheet.append(['Indicadores producción', 'Valor'])
    for cell in sheet[sheet.max_row]:
        cell.font = Font(bold=True)
    for key, value in produccion.items():
        sheet.append([key.replace('_', ' ').title(), value])

    sheet.append([])
    sheet.append(['Alertas'])
    for cell in sheet[sheet.max_row]:
        cell.font = Font(bold=True)
    for alerta in reporte.alertas or ['Sin alertas críticas']:
        sheet.append([alerta])

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def exportar_reporte_financiero_pdf(reporte: ReporteFinanciero) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    indicadores = reporte.indicadores or {}
    kpis = indicadores.get('kpis', {})
    produccion = indicadores.get('produccion', {})

    output = BytesIO()
    document = SimpleDocTemplate(output, pagesize=A4)
    styles = getSampleStyleSheet()
    story = [
        Paragraph('Reporte Financiero', styles['Title']),
        Paragraph(f'Periodo: {reporte.fecha_inicio} a {reporte.fecha_fin}', styles['BodyText']),
        Spacer(1, 12),
    ]

    table = Table(
        [['KPI', 'Valor']] + [[key.replace('_', ' ').title(), value] for key, value in kpis.items()],
        repeatRows=1,
    )
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f3e67')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#dce7f5')),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f8fbff')),
    ]))
    story.append(table)
    story.append(Spacer(1, 12))

    prod_table = Table(
        [['Indicador producción', 'Valor']] + [[key.replace('_', ' ').title(), value] for key, value in produccion.items()],
        repeatRows=1,
    )
    prod_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3f6b96')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#dce7f5')),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f8fbff')),
    ]))
    story.append(prod_table)
    story.append(Spacer(1, 12))
    story.append(Paragraph('Alertas', styles['Heading2']))
    for alerta in reporte.alertas or ['Sin alertas críticas']:
        story.append(Paragraph(f'- {alerta}', styles['BodyText']))

    document.build(story)
    return output.getvalue()


def generar_estado_financiero(usuario=None, fecha_inicio: date | None = None, fecha_fin: date | None = None, tipo: str = EstadoFinanciero.TipoEstado.RESULTADOS) -> EstadoFinanciero:
    data = calcular_dashboard_finanzas(fecha_inicio=fecha_inicio, fecha_fin=fecha_fin)
    kpis = data['kpis']
    nombre = f'{tipo.title()} {data["fecha_inicio"]} - {data["fecha_fin"]}'
    estado = EstadoFinanciero.objects.create(
        nombre=nombre,
        tipo=tipo,
        fecha_inicio=data['fecha_inicio'],
        fecha_fin=data['fecha_fin'],
        datos={
            'flujo_caja': float(kpis['flujo_caja']),
            'rentabilidad': float(kpis['rentabilidad']),
            'presupuesto_total': float(kpis['presupuesto_total']),
            'gasto_real': float(kpis['gasto_real']),
        },
        creado_por=usuario,
        actualizado_por=usuario,
    )
    return estado
